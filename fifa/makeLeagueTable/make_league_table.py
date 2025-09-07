import os
import time
import pandas as pd
import random
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


# ✅ 팀 리스트 정의
def get_js_teams():
    return ["Arsenal", "Newcastle", "Barcelona", "Inter Milan", "PSG"]


def get_hs_teams():
    return ["Manchester United", "Manchester City", "Real Madrid", "Liverpool", "Bayern Munich"]


def make_first_half(hs_teams, js_teams, num_rounds=5):
    """
    Junior 팀 vs Senior 팀만 매치 생성
    각 라운드에서 Junior 5팀 Senior 팀과 각각 경기. 총 5라운드
    """
    if len(js_teams) != len(hs_teams):
        raise ValueError("Junior 와 Senior 팀 수는 같아야 합니다.")

    n = len(js_teams)
    schedule = []

    # 라운드마다 senior 팀 순서를 회전시켜서 매치업 다양화
    for r in range(num_rounds):
        round_matches = []
        for i in range(n):
            home = js_teams[i]
            away = hs_teams[(i + r) % n]
            round_matches.append([home, "", "", away])
        schedule.append(round_matches)

    return schedule


def shuffle_each_round_order(schedule, start_round=1, end_round=5):
    """
    일정 중 특정 범위의 라운드들 안에서만 경기 순서를 섞는다.
    (예: 1~10R 사이의 각 라운드에서 순서만 랜덤하게 변경)
    """
    for i in range(start_round - 1, end_round):  # 0-based index
        random.shuffle(schedule[i])
    return schedule


def make_second_half(first_half):
    return [[ [match[3], "", "", match[0]] for match in round ] for round in first_half]


def shuffle_second_half_schedule(second_half):
    random.shuffle(second_half)
    for round_matches in second_half:
        random.shuffle(round_matches)
    return second_half


def check_duplicates(schedule, max_allowed=2):
    """
    경기 일정 검증 로직

    - 두 팀이 max_allowed(기본 2회)보다 많이 붙으면 경고
    - 총 경기 수 / 고유 매치업 수 출력
    """

    match_counter = defaultdict(int)
    match_rounds = defaultdict(list)

    for round_idx, round_matches in enumerate(schedule, 1):  # 1-based index
        for match in round_matches:
            key = tuple(sorted([match[0], match[3]]))  # 홈/원정 무시
            match_counter[key] += 1
            match_rounds[key].append(round_idx)

    # 중복 매치업 필터링
    true_duplicates = {k: v for k, v in match_counter.items() if v > max_allowed}

    if true_duplicates:
        print("\n🚨 중복 매치업 발견 (허용치 초과):")
        for (team1, team2), count in true_duplicates.items():
            rounds = match_rounds[(team1, team2)]
            rounds_str = ", ".join(f"{r}R" for r in rounds)
            print(f" - {team1} vs {team2}: {count}회 (라운드: {rounds_str})")
        print(f"❌ 중복된 조합 수: {len(true_duplicates)}")
        return False
    else:
        expected_matches = len(schedule) * len(schedule[0])
        actual_matches = sum(match_counter.values())
        print("✅ 모든 팀 조합이 정상 범위 내 등장")
        print(f"총 고유 매치업 수: {len(match_counter)}")
        print(f"총 경기 수: {actual_matches} (예상: {expected_matches})")
        return True


# ✅ 팀별 홈/원정 횟수 카운트 함수
def count_home_away_matches(schedule):
    stats = defaultdict(lambda: {"home": 0, "away": 0})
    for round_matches in schedule:
        for match in round_matches:
            home, away = match[0], match[3]
            stats[home]["home"] += 1
            stats[away]["away"] += 1
    return stats


# ✅ 통계 출력 함수
def print_home_away_summary(stats):
    print("\n📊 홈/원정 경기 수 요약:")
    print(f"{'팀명':<25} {'홈':>3}  {'원정':>4}  {'합계':>4}")
    print("-" * 40)
    for team, count in sorted(stats.items()):
        home = count["home"]
        away = count["away"]
        total = home + away
        print(f"{team:<25} {home:>3}   {away:>4}   {total:>4}")


# ✅ Excel 저장 함수
def save_schedule_to_excel(schedule, season_name):
    path = f"C:\\playground\\seasons\\{season_name}.xlsx"

    if os.path.exists(path):
        print(f"📄 기존 파일 발견: {path} → 삭제 시도 중...")
        if not try_remove_file(path):
            return

    for round_idx, matches in enumerate(schedule, 1):
        df = pd.DataFrame(matches, columns=["HOME", "", "", "AWAY"])
        index_labels = [f"{round_idx}R - {i + 1}번째 경기" for i in range(len(matches))]
        df.index = index_labels

        if not os.path.exists(path):
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=f"Round {round_idx}")
        else:
            with pd.ExcelWriter(path, mode='a', engine="openpyxl", if_sheet_exists="new") as writer:
                df.to_excel(writer, sheet_name=f"Round {round_idx}")
        print(f"✅ Round {round_idx} 저장 완료")

    adjust_all_sheets_column_width(path)


# ✅ 파일 삭제 시도 (파일 열려있을 경우 재시도)
def try_remove_file(path, retries=5, delay=5):
    for attempt in range(retries):
        try:
            if os.path.exists(path):
                os.remove(path)
                print(f"{path} 파일 삭제 성공")
                return True
            else:
                return True
        except PermissionError:
            print(f"파일이 열려있어 삭제 실패. {delay}초 후 재시도 중... ({attempt + 1}/{retries})")
            time.sleep(delay)
    print("❌ 파일 삭제 실패. 수동으로 닫은 후 다시 실행해 주세요.")
    return False


def adjust_all_sheets_column_width(path):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        for column_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(path)
    print("📏 모든 시트 열 너비 자동 조정 완료")


# ✅ 메인 실행 함수
def main():
    season_name = "2025_2"

    hs_teams = get_hs_teams()
    js_teams = get_js_teams()

    print("[1] 전반기 일정 생성 중...")
    first_half = make_first_half(js_teams, hs_teams, num_rounds=5)
    first_half = shuffle_each_round_order(first_half, start_round=1, end_round=5)

    print("[2] 후반기 일정 생성 중...")
    # first_half 를 뒤집기만 함
    second_half = make_second_half(first_half)
    second_half = shuffle_each_round_order(second_half, start_round=1, end_round=5)

    full_schedule = first_half + second_half

    print("[3] 중복 체크 중...")
    check_duplicates(full_schedule)

    print("[4] 홈/원정 통계 계산 중...")
    stats = count_home_away_matches(full_schedule)
    print_home_away_summary(stats)

    print("[5] Excel 저장 시작...")
    save_schedule_to_excel(full_schedule, season_name)

    print("🎉 모든 라운드 저장 완료!")


if __name__ == "__main__":
    main()
